from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response
from django.shortcuts import get_object_or_404
from django.http import HttpResponse, JsonResponse
from django.db.models import Sum, Q
from django.utils import timezone
from datetime import timedelta
import csv
import io
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from .models import TipoDocumento, Cliente, Compra
from .serializers import (
    TipoDocumentoSerializer,
    ClienteSerializer,
    ClienteBusquedaSerializer,
    CompraSerializer
)


class TipoDocumentoViewSet(viewsets.ReadOnlyModelViewSet):
    """ViewSet para consultar tipos de documento"""
    queryset = TipoDocumento.objects.filter(activo=True)
    serializer_class = TipoDocumentoSerializer


class ClienteViewSet(viewsets.ModelViewSet):
    """ViewSet para gestión de clientes"""
    queryset = Cliente.objects.select_related('tipo_documento').prefetch_related('compras')
    serializer_class = ClienteSerializer
    
    def get_serializer_class(self):
        if self.action == 'buscar':
            return ClienteBusquedaSerializer
        return ClienteSerializer
    
    @action(detail=False, methods=['get'])
    def buscar(self, request):
        """
        Busca un cliente por tipo y número de documento
        GET /api/clientes/buscar/?tipo_documento_id=1&numero_documento=123456789
        """
        tipo_documento_id = request.query_params.get('tipo_documento_id')
        numero_documento = request.query_params.get('numero_documento')
        
        if not tipo_documento_id or not numero_documento:
            return Response(
                {'error': 'Se requiere tipo_documento_id y numero_documento'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            from django.db.models import Prefetch
            cliente = Cliente.objects.select_related('tipo_documento').prefetch_related(
                Prefetch('compras', queryset=Compra.objects.order_by('-fecha_compra'))
            ).get(
                tipo_documento_id=tipo_documento_id,
                numero_documento=numero_documento,
                activo=True
            )
            serializer = self.get_serializer(cliente)
            return Response(serializer.data)
        except Cliente.DoesNotExist:
            return Response(
                {'error': 'Cliente no encontrado'},
                status=status.HTTP_404_NOT_FOUND
            )
    
    @action(detail=True, methods=['get'])
    def exportar(self, request, pk=None):
        """
        Exporta la información del cliente en diferentes formatos
        GET /api/clientes/{id}/exportar/?formato=csv|excel|txt
        """
        cliente = self.get_object()
        formato = request.query_params.get('formato', 'csv').lower()
        
        if formato == 'csv':
            return self._exportar_csv(cliente)
        elif formato == 'excel':
            return self._exportar_excel(cliente)
        elif formato == 'txt':
            return self._exportar_txt(cliente)
        else:
            return Response(
                {'error': 'Formato no válido. Use: csv, excel o txt'},
                status=status.HTTP_400_BAD_REQUEST
            )
    
    def _exportar_csv(self, cliente):
        """Exporta cliente a CSV"""
        response = HttpResponse(content_type='text/csv; charset=utf-8')
        response['Content-Disposition'] = f'attachment; filename="cliente_{cliente.numero_documento}.csv"'
        
        writer = csv.writer(response)
        writer.writerow(['Campo', 'Valor'])
        writer.writerow(['Tipo de Documento', cliente.tipo_documento.nombre])
        writer.writerow(['Número de Documento', cliente.numero_documento])
        writer.writerow(['Nombre', cliente.nombre])
        writer.writerow(['Apellido', cliente.apellido])
        writer.writerow(['Correo', cliente.correo])
        writer.writerow(['Teléfono', cliente.telefono])
        writer.writerow(['Fecha de Registro', cliente.fecha_registro.strftime('%Y-%m-%d %H:%M:%S')])
        
        # Agregar compras
        writer.writerow([])
        writer.writerow(['Compras'])
        writer.writerow(['Número Factura', 'Fecha', 'Monto', 'Estado'])
        for compra in cliente.compras.all():
            writer.writerow([
                compra.numero_factura,
                compra.fecha_compra.strftime('%Y-%m-%d'),
                f"${compra.monto:,.2f}",
                compra.estado
            ])
        
        return response
    
    def _exportar_excel(self, cliente):
        """Exporta cliente a Excel"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Cliente"
        
        # Estilos
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        # Datos del cliente
        ws.append(['Campo', 'Valor'])
        ws.append(['Tipo de Documento', cliente.tipo_documento.nombre])
        ws.append(['Número de Documento', cliente.numero_documento])
        ws.append(['Nombre', cliente.nombre])
        ws.append(['Apellido', cliente.apellido])
        ws.append(['Correo', cliente.correo])
        ws.append(['Teléfono', cliente.telefono])
        ws.append(['Fecha de Registro', cliente.fecha_registro.strftime('%Y-%m-%d %H:%M:%S')])
        
        # Compras
        ws.append([])
        ws.append(['Número Factura', 'Fecha', 'Monto', 'Estado'])
        header_row = ws[ws.max_row]
        for cell in header_row:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        
        for compra in cliente.compras.all():
            ws.append([
                compra.numero_factura,
                compra.fecha_compra.strftime('%Y-%m-%d'),
                compra.monto,
                compra.estado
            ])
        
        # Ajustar ancho de columnas
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 30
        
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="cliente_{cliente.numero_documento}.xlsx"'
        wb.save(response)
        return response
    
    def _exportar_txt(self, cliente):
        """Exporta cliente a TXT"""
        output = io.StringIO()
        output.write("=" * 50 + "\n")
        output.write("INFORMACIÓN DEL CLIENTE\n")
        output.write("=" * 50 + "\n\n")
        output.write(f"Tipo de Documento: {cliente.tipo_documento.nombre}\n")
        output.write(f"Número de Documento: {cliente.numero_documento}\n")
        output.write(f"Nombre: {cliente.nombre}\n")
        output.write(f"Apellido: {cliente.apellido}\n")
        output.write(f"Correo: {cliente.correo}\n")
        output.write(f"Teléfono: {cliente.telefono}\n")
        output.write(f"Fecha de Registro: {cliente.fecha_registro.strftime('%Y-%m-%d %H:%M:%S')}\n\n")
        
        output.write("=" * 50 + "\n")
        output.write("COMPRAS\n")
        output.write("=" * 50 + "\n\n")
        for compra in cliente.compras.all():
            output.write(f"Factura: {compra.numero_factura}\n")
            output.write(f"Fecha: {compra.fecha_compra.strftime('%Y-%m-%d')}\n")
            output.write(f"Monto: ${compra.monto:,.2f}\n")
            output.write(f"Estado: {compra.estado}\n")
            output.write("-" * 50 + "\n")
        
        response = HttpResponse(output.getvalue(), content_type='text/plain; charset=utf-8')
        response['Content-Disposition'] = f'attachment; filename="cliente_{cliente.numero_documento}.txt"'
        return response


class ReporteFidelizacionViewSet(viewsets.ViewSet):
    """ViewSet para generar reporte de fidelización de clientes"""
    
    @action(detail=False, methods=['get'])
    def generar(self, request):
        """
        Genera reporte en Excel de clientes elegibles para fidelización
        (compras > 5'000.000 COP en el último mes)
        GET /api/reporte-fidelizacion/generar/
        """
        # Calcular fecha del último mes
        fecha_limite = timezone.now() - timedelta(days=30)
        monto_minimo = 5000000  # 5 millones de pesos COP
        
        # Obtener clientes con compras en el último mes
        clientes = Cliente.objects.filter(
            compras__fecha_compra__gte=fecha_limite,
            compras__estado='completada',
            activo=True
        ).distinct()
        
        # Calcular total de compras por cliente en el último mes
        datos_reporte = []
        for cliente in clientes:
            total_compras = Compra.objects.filter(
                cliente=cliente,
                fecha_compra__gte=fecha_limite,
                estado='completada'
            ).aggregate(total=Sum('monto'))['total'] or 0
            
            if total_compras >= monto_minimo:
                datos_reporte.append({
                    'Tipo Documento': cliente.tipo_documento.nombre,
                    'Número Documento': cliente.numero_documento,
                    'Nombre': cliente.nombre,
                    'Apellido': cliente.apellido,
                    'Correo': cliente.correo,
                    'Teléfono': cliente.telefono,
                    'Total Compras (COP)': float(total_compras)
                })
        
        if not datos_reporte:
            return Response(
                {'mensaje': 'No hay clientes que cumplan los criterios de fidelización'},
                status=status.HTTP_404_NOT_FOUND
            )
        
        # Crear DataFrame con pandas
        df = pd.DataFrame(datos_reporte)
        df = df.sort_values('Total Compras (COP)', ascending=False)
        
        # Crear archivo Excel
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Clientes Fidelización')
            
            # Formatear la hoja
            worksheet = writer.sheets['Clientes Fidelización']
            
            # Estilos para encabezados
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center')
            
            # Ajustar ancho de columnas
            worksheet.column_dimensions['A'].width = 18
            worksheet.column_dimensions['B'].width = 20
            worksheet.column_dimensions['C'].width = 20
            worksheet.column_dimensions['D'].width = 20
            worksheet.column_dimensions['E'].width = 30
            worksheet.column_dimensions['F'].width = 15
            worksheet.column_dimensions['G'].width = 20
        
        output.seek(0)
        
        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        fecha_reporte = timezone.now().strftime('%Y%m%d_%H%M%S')
        response['Content-Disposition'] = f'attachment; filename="reporte_fidelizacion_{fecha_reporte}.xlsx"'
        return response

